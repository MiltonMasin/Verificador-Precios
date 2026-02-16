
from flask import Flask, send_from_directory, jsonify, request
import json
import os
from sync import parse_currency_to_number
import unicodedata

app = Flask(__name__, static_folder='static', template_folder='static')

DATA_PATH = os.path.join('data', 'productos.json')

@app.route('/')
def index():
    return send_from_directory('static', 'index.html')

@app.route('/productos')
def productos():
    try:
        with open(DATA_PATH, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return jsonify({"items": data})
    except Exception as e:
        return jsonify({"items": [], "error": str(e)}), 500

@app.route('/upload-excel', methods=['POST'])
def upload_excel():
    file = request.files.get('file')
    if not file:
        return jsonify({"error": "Archivo 'file' no enviado"}), 400
    try:
        try:
            from openpyxl import load_workbook
        except Exception as e:
            return jsonify({"error": f"Dependencia faltante para XLSX: {e}"}), 500

        def norm(s):
            s = str(s or '')
            s = unicodedata.normalize('NFD', s)
            s = ''.join(ch for ch in s if unicodedata.category(ch) != 'Mn')
            s = s.lower().strip().replace('\n', ' ').replace('\r', ' ')
            s = ' '.join(s.split())
            return s

        syn = {
            'codigo': {'codigo', 'código', 'cod', 'sku', 'id', 'codigo articulo', 'codigo producto', 'cod producto'},
            'nombre': {'nombre', 'descripcion', 'descripción', 'producto', 'detalle', 'articulo', 'artículo'},
            'precio': {
                'precio', 'precio venta', 'precio unitario', 'pvp', 'valor', 'importe', 'monto', 'precio lista',
                'p.venta', 'p venta', 'pventa', 'p.vta', 'pvta'
            }
        }

        wb = load_workbook(filename=file.stream, data_only=True)
        ws_candidate = None
        header_row_index = None
        indices = None
        for ws in wb.worksheets:
            max_check = min(ws.max_row, 10)
            for r in range(1, max_check + 1):
                row_cells = list(ws.iter_rows(min_row=r, max_row=r))[0]
                headers = [str(c.value or '').strip() for c in row_cells]
                nheaders = [norm(h) for h in headers]
                idx_map = {}
                for i, nh in enumerate(nheaders):
                    if nh in syn['codigo'] and 'Codigo' not in idx_map:
                        idx_map['Codigo'] = i
                    if nh in syn['nombre'] and 'Nombre' not in idx_map:
                        idx_map['Nombre'] = i
                    if nh in syn['precio'] and 'Precio' not in idx_map:
                        idx_map['Precio'] = i
                if all(k in idx_map for k in ['Codigo', 'Nombre', 'Precio']):
                    ws_candidate = ws
                    header_row_index = r
                    indices = idx_map
                    break
            if ws_candidate:
                break

        if not ws_candidate:
            return jsonify({"error": "No se encontraron columnas equivalentes a Codigo, Nombre, Precio"}), 400

        out = []
        for row in ws_candidate.iter_rows(min_row=header_row_index + 1):
            codigo = str(row[indices['Codigo']].value or '').strip()
            nombre = str(row[indices['Nombre']].value or '').strip()
            precio_raw = row[indices['Precio']].value
            if isinstance(precio_raw, (int, float)):
                precio_val = float(precio_raw)
            else:
                precio_str = str(precio_raw) if precio_raw is not None else ''
                precio_val = parse_currency_to_number(precio_str)
            if codigo and nombre:
                out.append({
                    'Codigo': codigo,
                    'Nombre': nombre,
                    'Precio': precio_val
                })

        os.makedirs(os.path.dirname(DATA_PATH), exist_ok=True)
        with open(DATA_PATH, 'w', encoding='utf-8') as f:
            json.dump(out, f, ensure_ascii=False, indent=2)

        return jsonify({"importados": len(out), "path": DATA_PATH})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/static/<path:filename>')
def static_files(filename):
    return send_from_directory('static', filename)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8080, debug=False)
